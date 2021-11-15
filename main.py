import openpyxl as xl

name_of_file = ''
wb = xl.load_workbook('Attendance Form Responses.xlsx')
sheet = wb['Form Responses 1']
names = {}

for row in range(3, sheet.max_row + 1):
    cell_coordinate = 'b' + str(row)
    cell = sheet[cell_coordinate]
    name = str(cell.value)
    shortened = name[0:-1]

    if name[-1] == ' ':
        name = shortened
    if name not in names:
        names[name] = 1
    else:
        names[name] += 1

del wb['Form Responses 1']
updated_sheet = wb.create_sheet('Points')
count = 1

for key in names:
    updated_cell_coordinate = 'a' + str(count)
    updated_cell = updated_sheet[updated_cell_coordinate]
    updated_cell.value = key

    updated_cell_coordinate = 'b' + str(count)
    updated_cell = updated_sheet[updated_cell_coordinate]
    updated_cell.value = names[key]

    count += 1

wb.save('Attendance Form With Points.xlsx')
